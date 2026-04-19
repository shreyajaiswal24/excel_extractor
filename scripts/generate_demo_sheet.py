"""Generate a demo Excel sheet for testing the Excel Extractor UI."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUT = Path(__file__).resolve().parent.parent / "public" / "demo-test-results.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADERS = [
    "Test Case ID",
    "Description",
    "Expected Result",
    "Actual Result",
    "Status",
    "Remarks",
]

MODULES = [
    ("Auth", [
        ("User logs in with valid credentials", "Dashboard loads", "Dashboard loads", "Pass", ""),
        ("User logs in with invalid password", "Show error message", "Generic 500 error", "Fail", "JIRA-1042"),
        ("Forgot password sends reset email", "Email arrives within 60s", "Email arrives in 12s", "Pass", ""),
        ("Account lockout after 5 failed attempts", "Account locked for 15m", "Not locked", "Defect Logged", "JIRA-1051"),
        ("OAuth login via Google", "Redirect to consent screen", "Pending env setup", "NA", "Blocked - staging not configured"),
    ]),
    ("Billing", [
        ("Create new subscription - monthly", "Charge processed, status active", "Charge processed", "Pass", ""),
        ("Apply discount coupon at checkout", "Total reduced by 20%", "Total reduced by 20%", "Pass", ""),
        ("Cancel subscription mid-cycle", "Refund prorated balance", "No refund issued", "Fail", "JIRA-1067"),
        ("Update payment method", "New card saved as default", "New card saved", "Pass", ""),
        ("Invoice PDF download", "PDF downloads with line items", "PDF empty", "Defect Logged", "JIRA-1071"),
        ("Tax calculation for EU customer", "VAT included", "Pending tax service", "NA", ""),
    ]),
    ("Reports", [
        ("Generate monthly sales report", "Report rendered with charts", "Report rendered", "Pass", ""),
        ("Export report as CSV", "CSV downloads with all rows", "CSV downloads", "Pass", ""),
        ("Filter report by date range", "Only matching rows shown", "All rows shown", "Fail", "JIRA-1082"),
        ("Schedule weekly email report", "Email sent every Monday 9am", "Sent at 11am", "Defect Logged", "JIRA-1085"),
        ("Drill-down on chart segment", "Detail view opens", "Detail view opens", "Pass", ""),
        ("Compare two periods side-by-side", "Two columns rendered", "Pending UI redesign", "NA", ""),
    ]),
    ("Search", [
        ("Search by exact keyword", "Top result matches keyword", "Top result matches", "Pass", ""),
        ("Fuzzy search with typo", "Suggested correction shown", "No suggestion", "Fail", "JIRA-1093"),
        ("Filter results by category", "Only category items shown", "Only category items shown", "Pass", ""),
        ("Search empty query", "Show recent searches", "Crashes app", "Defect Logged", "JIRA-1097"),
        ("Pagination on search results", "Page 2 loads next 20", "Page 2 loads next 20", "Pass", ""),
    ]),
    ("Profile", [
        ("Update display name", "Name reflected in header", "Name reflected", "Pass", ""),
        ("Upload avatar image", "Image cropped and saved", "Image cropped and saved", "Pass", ""),
        ("Delete account flow", "Confirmation modal appears", "Modal appears", "Pass", ""),
        ("Two-factor enrollment", "QR code shown, code accepted", "QR shown but code rejected", "Fail", "JIRA-1108"),
        ("Change email address", "Verification email sent", "Pending email service", "NA", ""),
    ]),
]


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    header_fill = PatternFill("solid", fgColor="0284C7")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="E0F2FE"),
        right=Side(style="thin", color="E0F2FE"),
        top=Side(style="thin", color="E0F2FE"),
        bottom=Side(style="thin", color="E0F2FE"),
    )
    status_fills = {
        "Pass": PatternFill("solid", fgColor="DCFCE7"),
        "Fail": PatternFill("solid", fgColor="FEE2E2"),
        "NA": PatternFill("solid", fgColor="F3F4F6"),
        "Defect Logged": PatternFill("solid", fgColor="FEF3C7"),
    }
    status_fonts = {
        "Pass": Font(color="166534", bold=True),
        "Fail": Font(color="991B1B", bold=True),
        "NA": Font(color="374151", bold=True),
        "Defect Logged": Font(color="92400E", bold=True),
    }

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    row_idx = 2
    tc_num = 1
    for module, cases in MODULES:
        for desc, expected, actual, status, remarks in cases:
            tc_id = f"TC-{tc_num:04d}"
            ws.append([tc_id, f"[{module}] {desc}", expected, actual, status, remarks])
            for col in range(1, 7):
                c = ws.cell(row=row_idx, column=col)
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=True)
            status_cell = ws.cell(row=row_idx, column=5)
            status_cell.fill = status_fills[status]
            status_cell.font = status_fonts[status]
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1
            tc_num += 1

    widths = [14, 50, 38, 38, 18, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote {OUT} ({tc_num - 1} test cases)")


if __name__ == "__main__":
    build()
